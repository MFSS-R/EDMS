"""
Excel import helpers for sample bulk import.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelImporter:
    """
    Create sample import templates and parse uploaded sample sheets.
    """

    HEADERS = [
        '所属项目',
        '实验',
        '样品类型',
        '显示代号',
        '样品名称',
        '批次号',
        '备注',
    ]

    @staticmethod
    def create_template():
        wb = Workbook()
        ws = wb.active
        ws.title = '样品导入模板'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for index, header in enumerate(ExcelImporter.HEADERS, start=1):
            cell = ws.cell(row=1, column=index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        widths = [18, 18, 16, 16, 20, 14, 14, 28]
        for column_index, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + column_index)].width = width

        ws.row_dimensions[1].height = 25
        ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def parse_uploaded_file(file):
        import pandas as pd

        df = pd.read_excel(file)
        required_columns = ['所属项目', '实验', '样品类型']
        for column in required_columns:
            if column not in df.columns:
                raise ValueError(f'缺少必填列: {column}')

        samples = []
        for index, row in df.iterrows():
            sample = {
                'project_name': str(row['所属项目']).strip() if pd.notna(row.get('所属项目', '')) else '',
                'experiment_name': str(row['实验']).strip() if pd.notna(row.get('实验', '')) else '',
                'sample_type_name': str(row['样品类型']).strip() if pd.notna(row.get('样品类型', '')) else '',
                'display_code': str(row['显示代号']).strip() if pd.notna(row.get('显示代号', '')) else '',
                'name': str(row['样品名称']).strip() if pd.notna(row.get('样品名称', '')) else '',
                'batch_number': str(row['批次号']).strip() if pd.notna(row.get('批次号', '')) else '',
                'notes': str(row['备注']).strip() if pd.notna(row.get('备注', '')) else '',
            }

            if not sample['project_name']:
                raise ValueError(f'第 {index + 2} 行：所属项目不能为空')
            if not sample['experiment_name']:
                raise ValueError(f'第 {index + 2} 行：实验不能为空')
            if not sample['sample_type_name']:
                raise ValueError(f'第 {index + 2} 行：样品类型不能为空')
            samples.append(sample)

        return samples

    @staticmethod
    def validate_and_create_samples(samples_data, default_project_id=None, user=None):
        from django.utils import timezone

        from apps.projects.models import Project
        from apps.samples.models import Experiment, Sample, SampleType

        created_samples = []
        errors = []
        created_projects = 0
        created_experiments = 0
        created_sample_types = 0

        project_cache = {}
        experiment_cache = {}
        sample_type_cache = {}

        for idx, sample_data in enumerate(samples_data):
            try:
                project_name = sample_data['project_name']
                experiment_name = sample_data['experiment_name']
                sample_type_name = sample_data['sample_type_name']

                if project_name not in project_cache:
                    project, created = Project.objects.get_or_create(
                        name=project_name,
                        user=user,
                        defaults={'description': '批量导入自动创建'},
                    )
                    project_cache[project_name] = project
                    if created:
                        created_projects += 1
                project = project_cache[project_name]

                experiment_key = f'{project.id}:{experiment_name}'
                if experiment_key not in experiment_cache:
                    experiment, created = Experiment.objects.get_or_create(
                        project=project,
                        name=experiment_name,
                        defaults={'description': '批量导入自动创建'},
                    )
                    experiment_cache[experiment_key] = experiment
                    if created:
                        created_experiments += 1
                experiment = experiment_cache[experiment_key]

                sample_type_key = f'{project.id}:{sample_type_name}'
                if sample_type_key not in sample_type_cache:
                    sample_type, created = SampleType.objects.get_or_create(
                        project=project,
                        name=sample_type_name,
                        defaults={'description': '批量导入自动创建'},
                    )
                    sample_type_cache[sample_type_key] = sample_type
                    if created:
                        created_sample_types += 1
                sample_type = sample_type_cache[sample_type_key]

                sample = Sample(
                    experiment=experiment,
                    sample_type=sample_type,
                    display_code=sample_data.get('display_code', ''),
                    name=sample_data.get('name', ''),
                    batch_number=sample_data.get('batch_number', ''),
                    notes=sample_data.get('notes', ''),
                    synthesis_date=timezone.now().date(),
                )
                sample.save()
                created_samples.append(sample)
            except Exception as exc:
                errors.append({
                    'row': idx + 2,
                    'message': str(exc),
                })

        return {
            'created_count': len(created_samples),
            'error_count': len(errors),
            'errors': errors,
            'created_projects': created_projects,
            'created_experiments': created_experiments,
            'created_sample_types': created_sample_types,
            'samples': [
                {
                    'sample_id': sample.sample_id,
                    'display_code': sample.display_code,
                    'name': sample.name,
                    'full_label': sample.full_label,
                    'sample_type_name': sample.sample_type.name,
                    'experiment_name': sample.experiment.name,
                    'project_name': sample.experiment.project.name,
                }
                for sample in created_samples
            ],
        }
