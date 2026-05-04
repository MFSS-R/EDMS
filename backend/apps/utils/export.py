"""
Excel export helpers.
"""

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Export sample and test data records to Excel workbooks.
    """

    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    @classmethod
    def export_samples(cls, samples):
        wb = Workbook()
        ws = wb.active
        ws.title = '样品列表'

        headers = [
            '系统编号',
            '显示代号',
            '样品名称',
            '完整标识',
            '所属项目',
            '所属实验',
            '样品类型',
            '合成日期',
            '批次号',
            '标记',
            '已测类型',
            '备注',
            '创建时间',
        ]

        for column, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=column, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cls.BORDER

        for row_index, sample in enumerate(samples, start=2):
            row = [
                sample.sample_id,
                sample.display_code,
                sample.name,
                sample.full_label,
                sample.project.name,
                sample.experiment.name if sample.experiment else '',
                sample.sample_type.name,
                sample.synthesis_date.strftime('%Y-%m-%d') if sample.synthesis_date else '',
                sample.batch_number,
                sample.mark,
                ', '.join(sample.test_types),
                sample.notes,
                sample.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]

            for column, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=column, value=value)
                cell.alignment = Alignment(vertical='center')
                cell.border = cls.BORDER

        widths = [22, 16, 20, 34, 18, 18, 16, 14, 14, 12, 24, 24, 20]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def export_test_data(cls, test_data_list):
        wb = Workbook()
        ws = wb.active
        ws.title = '测试数据'

        headers = [
            '系统编号',
            '显示代号',
            '样品名称',
            '完整标识',
            '测试类型',
            '测试日期',
            '测试仪器',
            '测试人员',
            '备注',
            '创建时间',
        ]

        for column, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=column, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cls.BORDER

        for row_index, test_data in enumerate(test_data_list, start=2):
            sample = test_data.sample
            row = [
                sample.sample_id,
                sample.display_code,
                sample.name,
                sample.full_label,
                test_data.test_type.name,
                test_data.test_date.strftime('%Y-%m-%d') if test_data.test_date else '',
                test_data.instrument,
                test_data.tester,
                test_data.notes,
                test_data.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]

            for column, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=column, value=value)
                cell.alignment = Alignment(vertical='center')
                cell.border = cls.BORDER

        widths = [22, 16, 20, 34, 18, 14, 18, 16, 24, 20]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def create_response(cls, buffer, filename):
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
